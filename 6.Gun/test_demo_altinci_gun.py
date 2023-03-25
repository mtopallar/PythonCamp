from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.action_chains import ActionChains
import pytest
from pathlib import Path
from datetime import date
import openpyxl
import sys
sys.path.append("./constants")
import globalConstants

#pytest --collect-only
class Test_DemoClass:
    #her testten önce çağırılır => her test öncesi ilgili test için genel bir setup gerekebilir.
    def setup_method(self):
        # mesela driver oluşturmak için veya pencerenin maximize edilmesi için kullanılabilir.
        # print("1")  print pytest eile test edilirken görüntülenmeyecektir. metodu boş bırakmamak için yazdık. breakpoint koyup debug edilebilir.
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.driver.maximize_window()
        self.driver.get(globalConstants.URL)
        #günün tarihini alıp bu tarih ile bir klasör var mı bak, yoksa oluştur.
        self.folderPath = str(date.today()) #yyyy-mm-dd
        Path(self.folderPath).mkdir(exist_ok=True) #exist_ok = True demek klasör zaten mevcutsa yeni oluşturma, mevcutu kabul et demek.
        
    
    #her testten sonra çağırılır => her test çalışmayı bitirdiğinde ilgili işlemleri temizlemek gibi bir durum gerekebilir.
    def teardown_method(self):
        #mesela her testte yeni taze bir chrome açılsın istersek driver ı kapatabiliriz. ya da mesela her test sonunda anasayfaya yönlendirme yapmak gerekbilir, onu da buradan yapabiliriz.
        # print("2")
        self.driver.quit() #!cookies
    
    def readData(self): #bu bir test fonksiyonu değil bir yardımcı fonksiyondur. onun için test kelimesi ile başlatmayız ve pytest bunu görmez.
        print("Excelden veri getirildi.")

    # Yani metod çalışırkenki sıralama: setup_method() => test_demoFunc() => teardown_method()
    def test_demoFunc(self):
        # 3A Act Arrange Assert
        # testin sonucunu bir condition a bağlamak için assert kullanılır. "Bütün bu olanlar sonucunda durum buysa testim doğrudur."
        
        text = "Hello"
        assert text == "Hello" # => bu blok sonucu test başarılı döner ancak;

        # text = "Hello"
        # assert text == "Hello 1" #=> bu testin başarısız olduğunu belirtir.
    
     # Yani metod çalışırkenki sıralama: setup_method() => test_demo2() => teardown_method()
    def test_demo2(self):
        assert True # => testi hep başarılı sayarken
        # assert False => testi hep başarısız sayar.
    
    def getData():
        #veriyi al
        # openpyxl python dan excel okumak, yazmak vs için pip paketi
        excelFile = openpyxl.load_workbook("./data/invalid_login.xlsx")
        selectedSheet = excelFile["Sayfa1"] #seçili sayfanın adı.

        totalRows = selectedSheet.max_row #içinde data olan satır sayısı (ilkini başlık saydık)
        data = []
        for i in range(2,totalRows + 1): #2 çünkü 1 başlık :)
            username = selectedSheet.cell(i,1).value
            password = selectedSheet.cell(i,2).value
            tupleData = (username,password)
            data.append(tupleData)

        return data
    
    #annotation - decorator
    
    @pytest.mark.parametrize("username,password",getData()) #d ecoratorden çalıştırılacak fonksiyona self demiyoruz
    def test_invalid_login(self,username,password):
        self.waifForElementVisible((By.ID,"user-name"))
        #WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,"user-name"))) 
        userNameInput = self.driver.find_element(By.ID,"user-name")
        self.waifForElementVisible((By.ID,"password"),10)
        #WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,"password"))) 
        passwordInput = self.driver.find_element(By.ID,"password")
        userNameInput.send_keys(username)
        passwordInput.send_keys(password)
        loginButton = self.driver.find_element(By.ID,"login-button")
        loginButton.click()
        errorMessage = self.driver.find_element(By.XPATH,"//*[@id='login_button_container']/div/form/div[3]/h3")
        self.driver.save_screenshot(f"{self.folderPath}/test_invalid_login_{username}_{password}.png")
        #magic string
        assert errorMessage.text == "Epic sadface: Username and password do not match any user in this service"
        
    def waifForElementVisible(self,locator,timeout = 5):
        WebDriverWait(self.driver,timeout).until(ec.visibility_of_element_located(locator))
# Annotation Notları:
# @pytest.mark.skip() #ilgili testi skip ettik skip(reason şeklinde de kullanılabilir.)
# @pytest.mark.parametrize()  ilgili testin "1" ve "1" değerleri için değil de farklı alternatifler ile de test edilmesini sağlar.
# @pytest.mark.parametrize("username,password",[("1","1"),("kullaniciAdim","sifrem")]) sırasıyla metoda userrname ve password olarak yine sırasıyla önce 1 , 1 sonra kullaniciAdim, sifrem değişkenlerini koyarak testi tekrar tekrar çalıştıracak. dictionary gibi. Bu şekilde bir test parametre sayısı kadar test ekranında da görünür (sol taraftaki alanda)